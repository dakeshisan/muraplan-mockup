"""Экспорт плана закупа в .xlsx (с CSV-фоллбэком) — лист «заказать», который
снабженец шлёт поставщикам. Зависимость необязательна: нет openpyxl → отдаём CSV.

Строки агрегируются по (объект, материал): сколько работ ждёт, к какой дате
заказать (самая ранняя), старт работ, запас в днях, статус. Только чтение.
"""
import io

STATUS_RU = {"overdue": "Просрочено", "now": "На этой неделе", "soon": "Скоро", "plan": "План"}
STATUS_RANK = {"overdue": 0, "now": 1, "soon": 2, "plan": 3}
HEADERS = ["Объект", "Материал", "Работ", "Заказать до", "Старт работ", "Запас, дн", "Статус"]


def buylist_rows(plan_data, statuses=("overdue", "now", "soon", "plan")):
    """Из плана «все объекты» (gpr_plan.summary_all) → агрегированный список к закупу."""
    agg = {}
    for it in (plan_data or {}).get("items", []):
        if it.get("status") not in statuses:
            continue
        k = (it.get("object"), it.get("material"))
        e = agg.get(k)
        if not e:
            agg[k] = {"object": it.get("object"), "material": it.get("material"), "count": 1,
                      "order_by": it["order_by"], "start": it.get("start"),
                      "slack": it.get("slack"), "status": it["status"]}
        else:
            e["count"] += 1
            e["order_by"] = min(e["order_by"], it["order_by"])
            if it.get("start") and (not e["start"] or it["start"] < e["start"]):
                e["start"] = it["start"]
            if it.get("slack") is not None and it["slack"] < e["slack"]:
                e["slack"] = it["slack"]
            if STATUS_RANK.get(it["status"], 9) < STATUS_RANK.get(e["status"], 9):
                e["status"] = it["status"]
    return sorted(agg.values(), key=lambda r: (r["order_by"], r["object"]))


def _fmt(s):
    if not s:
        return ""
    y, m, d = s.split("-")
    return f"{d}.{m}.{y}"


def to_csv(rows, today=None):
    import csv
    buf = io.StringIO()
    buf.write("﻿")                                  # BOM → Excel читает кириллицу
    w = csv.writer(buf, delimiter=";")
    w.writerow([f"ATLAS · план закупа — заказать (срез {today or ''})"])
    w.writerow(HEADERS)
    for r in rows:
        w.writerow([r["object"], r["material"], r["count"], _fmt(r["order_by"]),
                    _fmt(r["start"]), r["slack"], STATUS_RU.get(r["status"], r["status"])])
    return buf.getvalue().encode("utf-8"), "text/csv; charset=utf-8", f"atlas_zakup_{today or 'plan'}.csv"


def to_xlsx(rows, today=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception:
        return to_csv(rows, today)                       # нет openpyxl — отдаём CSV
    wb = Workbook()
    # детерминированные свойства — чтобы один и тот же план давал байт-в-байт один файл
    # (иначе openpyxl штампует время → авто-снимок Pages коммитит на каждый прогон)
    import datetime as _dt
    fixed = _dt.datetime(2026, 1, 1)
    wb.properties.creator = "ATLAS"
    wb.properties.created = fixed
    wb.properties.modified = fixed
    ws = wb.active
    ws.title = "Заказать"
    ws.append([f"ATLAS · план закупа — заказать (срез {today or ''})"])
    ws["A1"].font = Font(size=13, bold=True, color="FF284157")
    ws.append(HEADERS)
    for c in ws[2]:
        c.font = Font(size=10, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor="FF284157")
        c.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append([r["object"], r["material"], r["count"], _fmt(r["order_by"]),
                   _fmt(r["start"]), r["slack"], STATUS_RU.get(r["status"], r["status"])])
        cells = ws[ws.max_row]
        if r["status"] == "overdue":
            cells[6].font = Font(color="FFC7544A", bold=True)
        elif r["status"] == "now":
            cells[6].font = Font(color="FFC07A1E", bold=True)
    for col, wdt in zip("ABCDEFG", (16, 32, 8, 13, 13, 10, 16)):
        ws.column_dimensions[col].width = wdt
    ws.freeze_panes = "A3"
    bio = io.BytesIO()
    wb.save(bio)
    return (bio.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"atlas_zakup_{today or 'plan'}.xlsx")
