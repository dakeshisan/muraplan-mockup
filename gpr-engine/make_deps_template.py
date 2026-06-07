#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ATLAS · ГПР-движок — генератор шаблона связей зависимостей.

Делает gpr_dependencies.xlsx: лист на каждый ЖК, все работы с уникальным UID и
колонкой «Предшественники(UID)», ПРЕДЗАПОЛНЕННОЙ авто-оценкой (finish-to-start
по датам). Пользователь правит → возвращает → import_deps.py вшивает реальные
связи в gpr_data.json (поле deps) для настоящих стрелок и критического пути.

UID уникален в пределах листа (ЖК). Предшественники — UID из того же блока.
"""
import json, os, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = json.load(open(os.path.join(HERE, "gpr_data.json"), encoding="utf-8"))
DAY = 86400000


def dnum(s):
    d = datetime.date.fromisoformat(s)
    return (d - datetime.date(1970, 1, 1)).days


def infer_pred(works):
    """FS-оценка: для каждой работы — UID-позиция предшественника в блоке (или None)."""
    nd = [(i, dnum(w["start"]), dnum(w["end"])) for i, w in enumerate(works)]
    pred = [None] * len(works)
    for i, s, e in nd:
        best = None
        for j, sj, ej in nd:
            if j != i and ej <= s + 4 and ej >= s - 60 and (best is None or ej > best[2]):
                best = (j, sj, ej)
        if best and (s - best[2]) <= 21:
            pred[i] = best[0]  # локальный индекс в блоке
    return pred


# стили
NAVY = "284157"; TEAL = "007484"; GOLD = "CFB372"; CREAM = "F7F5EE"
hdr_fill = PatternFill("solid", fgColor=NAVY)
hdr_font = Font(color="FFFFFF", bold=True, size=11)
edit_fill = PatternFill("solid", fgColor="FFF3C4")      # жёлтая — заполнять
edit_hdr = PatternFill("solid", fgColor="C9A227")
shade_a = PatternFill("solid", fgColor="FFFFFF")
shade_b = PatternFill("solid", fgColor="EFEDE5")
crit_font = Font(color="C8554C", bold=True)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical="center", wrap_text=True)
center = Alignment(horizontal="center", vertical="center")

COLS = ["UID", "Блок", "№", "Раздел", "Работа", "Начало", "Окончание",
        "Дни", "Отст,дн", "Предшественники (UID)", "Тип", "Примечание"]
WIDTHS = [6, 9, 5, 16, 40, 11, 11, 6, 8, 22, 7, 24]
EDIT_COL = 10  # «Предшественники»

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── лист-инструкция ──
ins = wb.create_sheet("📖 Инструкция")
ins.sheet_view.showGridLines = False
ins.column_dimensions["A"].width = 110
lines = [
    ("ATLAS · ГПР-движок — связи зависимостей работ", True),
    ("", False),
    ("Зачем: задать реальные связи «какая работа должна завершиться перед началом другой».", False),
    ("Движок нарисует настоящие стрелки на Гантте и посчитает истинный критический путь.", False),
    ("", False),
    ("Как заполнять:", True),
    ("1) Один лист = один ЖК. В колонке «Предшественники (UID)» (жёлтая) укажите UID работ,", False),
    ("   которые должны завершиться ДО начала этой — через запятую (напр.: 3 или 3,5).", False),
    ("2) UID берите из ЭТОГО ЖЕ блока (см. колонку «Блок»). UID уникальны в пределах листа.", False),
    ("3) Колонка УЖЕ предзаполнена авто-оценкой по датам — просто ПОПРАВЬТЕ где неверно,", False),
    ("   пустые добавьте, лишние удалите. Если предшественника нет — оставьте пусто.", False),
    ("4) «Тип» обычно FS (finish→start). SS/FF/SF — если нужно (можно не трогать).", False),
    ("5) Можно заполнить хоть один блок — движок применит реальные связи там, где заданы,", False),
    ("   и оставит авто-оценку для остального.", False),
    ("", False),
    ("Сохраните как .xlsx и пришлите обратно. ATLAS · ATAMŪRA GROUP.", False),
]
for i, (t, b) in enumerate(lines, 1):
    c = ins.cell(row=i, column=1, value=t)
    c.font = Font(bold=b, size=14 if (b and i == 1) else 11,
                  color=NAVY if b else "1E1E1E")

# ── листы ЖК ──
for o in DATA["portfolio"]:
    title = o["object"].replace("ЖК ", "").replace("«", "").replace("»", "")[:28]
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    # заголовок
    for ci, name in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.fill = edit_hdr if ci == EDIT_COL else hdr_fill
        c.font = hdr_font; c.alignment = center; c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS[ci - 1]
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    uid = 0
    r = 2
    for bi, b in enumerate(o["blocks"]):
        pred = infer_pred(b["works"])
        base_uid = uid + 1
        uid += len(b["works"])
        for li, w in enumerate(b["works"]):
            myuid = base_uid + li
            puid = "" if pred[li] is None else str(base_uid + pred[li])
            vals = [myuid, "«%s»" % b["spot"], w["no"], w.get("section", ""),
                    w["name"], w["start"], w["end"], w.get("days", ""),
                    w.get("lag", 0), puid, "FS", w.get("contractor", "") or ""]
            shade = shade_a if (bi % 2 == 0) else shade_b
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=ci, value=v)
                c.border = border
                c.alignment = wrap if ci in (4, 5, 12) else center
                c.fill = edit_fill if ci == EDIT_COL else shade
                if ci == 9 and (w.get("lag", 0) or 0) > 0:
                    c.font = crit_font
            r += 1

dest = os.path.join(HERE, "gpr_dependencies.xlsx")
wb.save(dest)
nb = sum(len(o["blocks"]) for o in DATA["portfolio"])
nw = sum(len(b["works"]) for o in DATA["portfolio"] for b in o["blocks"])
print(f"✓ {dest}\n  листов {len(DATA['portfolio'])} (ЖК) · блоков {nb} · работ {nw}")
print("  колонка J «Предшественники» предзаполнена авто-оценкой (FS по датам).")
