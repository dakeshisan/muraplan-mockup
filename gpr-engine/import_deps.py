#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ATLAS · ГПР-движок — импорт связей зависимостей из заполненного xlsx.

Читает gpr_dependencies.xlsx (колонка «Предшественники (UID)») и вшивает в
gpr_data.json поле deps у работ: deps=[{b:блок_idx, w:работа_idx, t:тип}].
После — parse_gpr.py --build, и движок рисует НАСТОЯЩИЕ стрелки + критический путь.

UID реконструируются в том же порядке, что в make_deps_template.py
(объект → блоки по порядку → работы по порядку, сквозной счётчик на лист).
"""
import json, os, sys
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_PATH = os.path.join(HERE, "gpr_data.json")
XLSX = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "gpr_dependencies.xlsx")


def sheet_title(o):
    return o["object"].replace("ЖК ", "").replace("«", "").replace("»", "")[:28]


def main():
    data = json.load(open(DATA_PATH, encoding="utf-8"))
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    total = 0
    for o in data["portfolio"]:
        title = sheet_title(o)
        if title not in wb.sheetnames:
            print(f"  ! лист «{title}» не найден — пропуск"); continue
        ws = wb[title]
        # UID -> (block_idx, work_idx)
        uid2pos, uid = {}, 0
        for bi, b in enumerate(o["blocks"]):
            for wi in range(len(b["works"])):
                uid += 1; uid2pos[uid] = (bi, wi)
        # строки идут в том же порядке (строка 2 = uid 1)
        r, cnt = 2, 0
        for bi, b in enumerate(o["blocks"]):
            for wi, w in enumerate(b["works"]):
                pre = ws.cell(r, 10).value
                typ = (str(ws.cell(r, 11).value or "FS")).upper().strip()
                deps = []
                if pre not in (None, ""):
                    for tok in str(pre).replace(";", ",").replace(" ", ",").split(","):
                        tok = tok.strip()
                        if tok.isdigit() and int(tok) in uid2pos:
                            pb, pw = uid2pos[int(tok)]
                            if not (pb == bi and pw == wi):           # не сам на себя
                                deps.append({"b": pb, "w": pw, "t": typ})
                if deps:
                    w["deps"] = deps; cnt += 1
                else:
                    w.pop("deps", None)
                r += 1
        print(f"  {o['id']:7} связей задано у {cnt} работ")
        total += cnt
    data["deps_source"] = os.path.basename(XLSX)
    json.dump(data, open(DATA_PATH, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"✓ {DATA_PATH} обновлён · всего работ со связями: {total}")
    print("  далее: python3 parse_gpr.py --build  → движок нарисует реальные стрелки")


if __name__ == "__main__":
    main()
